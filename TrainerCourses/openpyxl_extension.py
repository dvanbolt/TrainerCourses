from __future__ import annotations
import re
import openpyxl
from typing import Optional,Any,Generator,Callable,Iterator,Iterable,NamedTuple
from pathlib import Path
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
from collections import namedtuple
from dataclasses import dataclass,field
from types import MethodType

_REPCHAR = ['/','-',' ','\\','&',]
_REPCHAR_RE = re.compile(r'|'.join([fr"{c}+" for c in _REPCHAR]))
_SNK_RE = re.compile(r'(?<!^)(?=[A-Z])(\s)')

class Printable:
    def __str__(self)->str:
        return f"{self.__class__.__name__}({', '.join([k+'='+str(v) for (k,v) in self.__dict__.items()])})"

@dataclass
class ImplicitNamedRange:
    name:str
    min_row:int   
    min_col:int
    max_col:int
    max_row:int
    sheet:Worksheet = field(repr=False)
    _nested:list[tuple[int,int]] = field(default_factory=list,repr=False)
    _nt:type = field(repr=False,default=None)

    def implicit_named_ranges(self)->dict[tuple[int,str],ImplicitNamedRange]:
        ret = {}
        if self._nested:
            for t in self._nested:
                nr = self.sheet._nr[t]
                ret[t] = nr
                for k,v in nr.implicit_named_ranges().items():
                    ret[k] = v                
        return ret

    @property
    def has_nested(self)->bool:
        return self._nested != []

    @staticmethod
    def _var_name(o:Any)->str:
        o = _REPCHAR_RE.sub('_',o)
        o = "".join(c for c in str(o) if c.isalnum() or c=='_')
        if not o or o[0].isnumeric():
            o = '_'+o
        return o


    @classmethod
    def _snake_case(cls,o:str)->str:
        return _SNK_RE.sub('_', cls._var_name(o)).lower()

    @classmethod
    def _camel_case(cls,o:str)->str:
        return ''.join(w.title() for w in cls._snake_case(o).split('_'))

    @property
    def nrows(self)->int:
        return self.max_row-self.min_row

    @property
    def top_row(self)->list[Any]:
        return list(self.iter_rows(max_row=1,values_only=True))[0]

    @property
    def header(self)->list[str]:
        return [self._snake_case(str(c)) for c in self.top_row]

    def iter_rows(self,values_only:bool=False,min_row:int=0,max_row:int=0)->Generator:
        min_row = self.min_row+min_row
        if not max_row or min_row+max_row>self.max_row:
            max_row = self.max_row
        else:
            max_row = min_row + max_row

      
        return self.sheet.iter_rows(min_row=min_row,
                                    max_row=max_row,
                                    max_col=self.max_col,
                                    min_col=self.min_col,
                                    values_only=values_only)


    def named_tuples(self,)->Generator[NamedTuple,None,None]:
        if not self._nt:
            try:
                self._nt = namedtuple(self._camel_case(self.name),' '.join(self.header))
            except ValueError as ve:
                raise Exception("Invalid header for namedtuple type",ve)
        for row in self.iter_rows(min_row=1,values_only=True):

            yield self._nt(*row)

    def object(self,key_index:int=0,value_index:int=1):
        obj = type(self._camel_case(self.name), (Printable,), {})()
        for row in self.iter_rows(min_row=0,values_only=True):
            setattr(obj,self._snake_case(row[key_index]), row[value_index])
        return obj

    def dict(self,key_index:int=0,value_index:int=1,snake_case_keys:bool=False)->dict:
        if snake_case_keys:
            return {self._snake_case(row[key_index]): row[value_index] for \
                    row in self.iter_rows(min_row=0,values_only=True)}
        else:
            return {row[key_index]: row[value_index] for \
                    row in self.iter_rows(min_row=0,values_only=True)}
    def list(self,element:type=dict,snake_case_keys:bool=False,values_only:bool=True,
             element_keys:list[int]|list[str]=None)->list[tuple]|list[dict]:
        ret = []
        key_filter = lambda x,e:x is not None
        if element_keys:
            if isinstance(element_keys[0],str):
                key_filter = lambda x,e:x in e
        if element==dict:
            if snake_case_keys:
                header = self.header
            else:
                header = self.top_row
            _hi = []
            _hd = []
            for c,h in enumerate(header):
                if key_filter(h,element_keys):
                    _hi.append(c)
            for row in self.iter_rows(min_row=1,values_only=values_only):
                nr = {}
                for i in _hi:
                    nr[header[i]] = row[i]
                if any(nr.values()):
                    ret.append(nr)
                else:
                    break
        elif element in (tuple,list):
            for row in self.iter_rows(min_row=0,values_only=values_only):
                ret.append(element(row))
        return ret

def _set_sheet_inrs(sheet:Worksheet)->None:
    last_merge_row = max(sheet._merged_cells.keys(),default=0)
    for row,col_sets in list(sheet._merged_cells.items()):
        for start,end in col_sets:
            title = sheet.cell(row=row+1,column=start+1).value
            height = 0
            for sub_row in sheet.iter_rows(min_row=row+1,min_col=start+1,max_col=end+1,values_only=True):
                if any(sub_row):
                    height+=1
                else:
                    break
            sr = ImplicitNamedRange(title,max_row=row+height,min_row=row+2,min_col=start+1,max_col=end+1,sheet=sheet)
            if row < last_merge_row:
                for next_start,next_end in sheet._merged_cells[row+1]:
                    if next_start<start:
                        continue
                    elif next_start>end:
                        break
                    else:
                        sr._nested.append((sheet.cell(row=row+2,column=next_start+1).value,row+1))
            suffix=1
            while (title,row) in sheet._nr:
                title = title + f'-{suffix}'
                suffix+=1
                if suffix>100:
                    raise ValueError((title,row))
            sheet._nr[(title,row)] = sr

def implicit_named_ranges(sheet:Worksheet)->dict[tuple[str,int],ImplicitNamedRange]:
    if not sheet._nr:
        _set_sheet_inrs(sheet)
    return sheet._nr


 
def clear_values(self):
    for row in self:
        for cell in row:
            cell.value = None

def write_row(self,row:list,offset:tuple[int,int]=(1,1))->None:
    for column_count,value in enumerate(row):
        self.cell(row=offset[0],column=offset[1]+column_count).value = value

def write_rows(self,rows:list[list]|list[dict],offset:tuple[int,int]=(1,1))->None:    
    if rows and isinstance(rows[0],dict):
        row_values = [list(rows[0].keys())]
        for row in rows:
            row_values.append(list(row.values()))
    for row_count,val in enumerate(row_values):
        self.write_row(val,(row_count+offset[0],offset[1]))

def open(pth:Path|str)->list[Course]:
    def _blanks(v:str|float|int):
        if v == "":
            return None
        return v
    pth = Path(pth)
    if pth.exists():
        wb = openpyxl.open(pth)
        for sc,sheet in enumerate(wb):
            sheet._merged_cells = {}
            sheet._nr = {}
            sheet.write_row = MethodType(write_row,sheet)
            sheet.write_rows = MethodType(write_rows,sheet)
            sheet.clear_values = MethodType(clear_values,sheet)
            if not sc:
                type(sheet).implicit_named_ranges = implicit_named_ranges
            
            for rc,row in enumerate(sheet.iter_rows()):
                row_keys = []
                key = []
                if any(isinstance(cell, openpyxl.cell.cell.MergedCell) for cell in row):
                    for cc,cell in enumerate(row):
                        if not isinstance(cell,openpyxl.cell.cell.MergedCell):
                            if len(key)>1:
                                row_keys.append(tuple([key[0],key[-1]]))
                            key = [cc]                           
                        else:
                            key.append(cc)
                    if len(key)>1:
                        row_keys.append(tuple([key[0],key[-1]]))
                    sheet._merged_cells[rc] = list(row_keys)
                else:                
                    break
       

        return wb
            
            
            
            


#wb = open(r'test\ATS.xlsx')
#for sheet in wb:
    #print(sheet.title)
    #print(list(sheet.implicit_named_ranges()))
    